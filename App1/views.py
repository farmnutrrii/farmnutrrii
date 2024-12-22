from datetime import datetime
from django.shortcuts import render, redirect
from django.http import HttpResponse
import openpyxl
from .models import ClosingStock, DailySheet, DailySales, WeeklyReport
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.contrib import messages
from decimal import Decimal
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.views.decorators.cache import cache_control
from django.contrib.auth.decorators import login_required

# Create your views here.
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)
            return redirect('index')  
        else:
            messages.error(request, 'Invalid credentials')
            return redirect('login')
    else:
        return render(request, 'login.html')

def logout_view(request):
    auth_logout(request)
    return redirect('login')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def index(request):

    return render(request, 'index.html')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def closing_stock_view(request):
    if request.method == 'POST':
        if 'delete' in request.POST:
            report_date = request.POST.get('report_date')
            try:
                report_date = datetime.strptime(report_date, "%Y-%m-%d").date()
                ClosingStock.objects.filter(date=report_date).delete()
            except ValueError:
                return redirect('closing_stock')

            return redirect('closing_stock')
        
        date = request.POST.get('date')
        day = request.POST.get('day')

        # Handle Broiler
        broiler_no_of_birds = request.POST.get('broiler_no_of_birds')
        broiler_no_of_kgs = request.POST.get('broiler_no_of_kgs')
        broiler_mortality = request.POST.get('broiler_mortality')

        if broiler_no_of_birds or broiler_no_of_kgs or broiler_mortality:
            ClosingStock.objects.create(
                date=date,
                day=day,
                bird_type='BROILER',
                no_of_birds=int(broiler_no_of_birds) if broiler_no_of_birds else 0,
                no_of_kgs=Decimal(broiler_no_of_kgs) if broiler_no_of_kgs else Decimal('0.00'),
                mortality=int(broiler_mortality) if broiler_mortality else 0
            )

        # Handle CC
        cc_no_of_birds = request.POST.get('cc_no_of_birds')
        cc_no_of_kgs = request.POST.get('cc_no_of_kgs')
        cc_mortality = request.POST.get('cc_mortality')

        if cc_no_of_birds or cc_no_of_kgs or cc_mortality:
            ClosingStock.objects.create(
                date=date,
                day=day,
                bird_type='CC',
                no_of_birds=int(cc_no_of_birds) if cc_no_of_birds else 0,
                no_of_kgs=Decimal(cc_no_of_kgs) if cc_no_of_kgs else Decimal('0.00'),
                mortality=int(cc_mortality) if cc_mortality else 0
            )

        # Handle Original
        original_no_of_birds = request.POST.get('original_no_of_birds')
        original_no_of_kgs = request.POST.get('original_no_of_kgs')
        original_mortality = request.POST.get('original_mortality')

        if original_no_of_birds or original_no_of_kgs or original_mortality:
            ClosingStock.objects.create(
                date=date,
                day=day,
                bird_type='ORIGINAL',
                no_of_birds=int(original_no_of_birds) if original_no_of_birds else 0,
                no_of_kgs=Decimal(original_no_of_kgs) if original_no_of_kgs else Decimal('0.00'),
                mortality=int(original_mortality) if original_mortality else 0
            )

        # Handle Quail
        quail_no_of_birds = request.POST.get('quail_no_of_birds')
        quail_no_of_kgs = request.POST.get('quail_no_of_kgs')
        quail_mortality = request.POST.get('quail_mortality')

        if quail_no_of_birds or quail_no_of_kgs or quail_mortality:
            ClosingStock.objects.create(
                date=date,
                day=day,
                bird_type='QUAIL',
                no_of_birds=int(quail_no_of_birds) if quail_no_of_birds else 0,
                no_of_kgs=Decimal(quail_no_of_kgs) if quail_no_of_kgs else Decimal('0.00'),
                mortality=int(quail_mortality) if quail_mortality else 0
            )

        return redirect('closing_stock')  # Redirect after saving

    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Retrieve all distinct date and day combinations
    stocks = ClosingStock.objects.values('date', 'day').distinct().order_by('-date')

    if date_filter:
        stocks = stocks.filter(date=date_filter)
    elif start_date and end_date:
        stocks = stocks.filter(date__range=[start_date, end_date])

    return render(request, 'closing_stock.html', {'stocks': stocks, 'date_filter': date_filter, 'start_date': start_date, 'end_date': end_date})

def view_stock(request, stock_date):
    # Query all stock entries for the given date and order by bird_type
    stocks = ClosingStock.objects.filter(date=stock_date).order_by('bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given date
        return HttpResponse("No records found for the given date.", status=404)

    # Get the day from the first stock entry (since all entries will have the same date and day)
    stock_day = stocks.first().day

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closing Report"

    ws.merge_cells('A1:E1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add the "Closing Stock" heading, center it, and apply bold styling
    ws.merge_cells('A2:E2')  # Merge cells to center the heading
    ws['A2'] = "Closing Stock"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add the date and day, and make the heading bold
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stock_day])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers, add a "S.NO" column, and make the headers bold
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "NO.OF.KGS", "MORTALITY"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables to track the current group and serial number
    current_bird_type = None
    serial_number = 1

    for stock in stocks:
        if stock.bird_type != current_bird_type:
            # Start a new group if the bird_type changes
            current_bird_type = stock.bird_type

        # Append stock data for the current group, including the serial number
        row = [
            serial_number,
            stock.bird_type,
            stock.no_of_birds,
            stock.no_of_kgs,
            stock.mortality,
        ]
        ws.append(row)

        # Apply bold styling to bird_type and center alignment to all cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:  # Make bird_type (column 2) bold
                cell.font = Font(bold=True)

        serial_number += 1

    # Set column widths for better visibility
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Closing_Report_{stock_date}.xlsx'

    wb.save(response)
    return response

def download_excel_closingstock(request):
    # Get the date filter from the query parameters
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter stocks based on the date_filter if it is provided
    if date_filter:
        stocks = ClosingStock.objects.filter(date=date_filter).order_by('bird_type')
    elif start_date and end_date:
        stocks = ClosingStock.objects.filter(date__range=[start_date, end_date]).order_by('date', 'bird_type')
    else:
        stocks = ClosingStock.objects.all().order_by('date', 'bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given filter
        return HttpResponse("No records found for the given date range.", status=404)

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closing Stock"

    # Add the main heading
    ws.merge_cells('A1:H1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')  # Merge cells to center the heading
    ws['A2'] = "Weekly Report"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add a blank row and set the headers for date and day columns
    ws.append([""])
    ws.append(["Date:", "", "", "Day:", ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers and make the headers bold
    headers = ["DATE", "DAY", "BIRDS TYPE", "NO.OF.BIRDS", "NO.OF.KGS", "MORTALITY"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables for tracking merging and row positions
    previous_date = None
    row_start = 7  # Row number where the data starts

    # Add stock data to the Excel sheet with merged date and day cells
    for stock in stocks:
        current_row = ws.max_row + 1

        # Insert a gap row if the current stock's date is different from the previous date
        if stock.date != previous_date and previous_date is not None:
            ws.append([""] * len(headers))  # Insert an empty row
            current_row += 1

        row = [
            stock.date if stock.date != previous_date else "",  # Only show date if it's different from the previous row
            stock.day if stock.date != previous_date else "",  # Only show day if it's different from the previous row
            stock.bird_type,
            stock.no_of_birds,
            stock.no_of_kgs,
            stock.mortality,
        ]
        ws.append(row)

        # Center align the date and day columns
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Merge date and day cells if necessary
        if stock.date != previous_date:
            if current_row > row_start:
                ws.merge_cells(start_row=row_start, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=row_start, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')
            row_start = current_row
            previous_date = stock.date

        # Apply styling to the rest of the row cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 3:  # Make bird_type (column 3) bold
                cell.font = Font(bold=True)

    # Merge the last set of date and day cells
    if row_start < ws.max_row:
        ws.merge_cells(start_row=row_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=row_start, start_column=2, end_row=ws.max_row, end_column=2)
        ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set up the filename
    if date_filter:
        filename = f"Closing_Stock_{date_filter}.xlsx"
    elif start_date and end_date:
        filename = f"Closing_Stock_{start_date}_to_{end_date}.xlsx"
    else:
        filename = "Closing_Stock.xlsx"
    
    # Set up the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def dailysheet(request):
    if request.method == 'POST':
        if 'delete' in request.POST:
            report_date = request.POST.get('report_date')
            print(f"Report date received: {report_date}")  # Debugging
            try:
                # Convert to date object
                report_date = datetime.strptime(report_date.strip(), "%Y-%m-%d").date()
                # Delete matching records
                deleted_count, _ = DailySheet.objects.filter(date=report_date).delete()
                if deleted_count == 0:
                    print(f"No records found for date: {report_date}")  # Debugging
            except ValueError as e:
                return redirect('dailysheet')

            return redirect('dailysheet')
        
        date = request.POST.get('date')
        day = request.POST.get('day')

        bird_types = ['broiler', 'cc', 'original', 'quail']
        
        for bird_type in bird_types:
            number_of_birds_stock = request.POST.get(f'{bird_type}_number_of_birds_stock', 0)
            number_of_birds_purchase = request.POST.get(f'{bird_type}_number_of_birds_purchase', 0.0)
            total_birds = request.POST.get(f'{bird_type}_total_birds', 0)
            total_stock_weight = request.POST.get(f'{bird_type}_total_stock_weight', 0.0)
            total_purchase_weight = request.POST.get(f'{bird_type}_total_purchase_weight', 0.0)
            total_weight = request.POST.get(f'{bird_type}_total_weight', 0.0)
            
            # Only create record if any relevant field is provided
            if (number_of_birds_stock or number_of_birds_purchase or total_birds or 
                total_stock_weight or total_purchase_weight or total_weight):
                DailySheet.objects.create(
                    date=date,
                    day=day,
                    bird_type=bird_type.capitalize(),
                    number_of_birds_stock=int(number_of_birds_stock),
                    number_of_birds_purchase=Decimal(number_of_birds_purchase),
                    total_birds=int(total_birds),
                    total_stock_weight=Decimal(total_stock_weight),
                    total_purchase_weight=Decimal(total_purchase_weight),
                    total_weight=Decimal(total_weight)
                )

        return redirect('dailysheet')  # Redirect after saving

    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Retrieve all distinct date and day combinations
    stocks = DailySheet.objects.values('date', 'day').distinct().order_by('-date')

    if date_filter:
        stocks = stocks.filter(date=date_filter)
    elif start_date and end_date:
        stocks = stocks.filter(date__range=[start_date, end_date])
    
    return render(request, 'dailysheet.html', {'stocks': stocks, 'date_filter': date_filter, 'start_date': start_date, 'end_date': end_date})

def view_stock_dailysheet(request, stock_date):
    # Query all stock entries for the given date and order by bird_type
    stocks = DailySheet.objects.filter(date=stock_date).order_by('bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given date
        return HttpResponse("No records found for the given date.", status=404)

    # Get the day from the first stock entry (since all entries will have the same date and day)
    stock_day = stocks.first().day

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Stock"

    ws.merge_cells('A1:E1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add the "Closing Stock" heading, center it, and apply bold styling
    ws.merge_cells('A2:E2')  # Merge cells to center the heading
    ws['A2'] = "Opening Stock"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add the date and day, and make the heading bold
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stock_day])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers, add a "S.NO" column, and make the headers bold
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS  C/O STOCK", "NO.OF.BIRS PURCHASE", "TOTAL BIRDS", "TOTAL C/O STOCK WEIGHT", "TOTAL PURCHASE WEIGHT", "TOTAL WEIGHT"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables to track the current group and serial number
    current_bird_type = None
    serial_number = 1

    for stock in stocks:
        if stock.bird_type != current_bird_type:
            # Start a new group if the bird_type changes
            current_bird_type = stock.bird_type

        # Append stock data for the current group, including the serial number
        row = [
            serial_number,
            stock.bird_type,
            stock.number_of_birds_stock,
            stock.number_of_birds_purchase,
            stock.total_birds,
            stock.total_stock_weight,
            stock.total_purchase_weight,
            stock.total_weight
        ]
        ws.append(row)

        # Apply bold styling to bird_type and center alignment to all cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:  # Make bird_type (column 2) bold
                cell.font = Font(bold=True)

        serial_number += 1

    # Set column widths for better visibility
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Opening_Stock_{stock_date}.xlsx'

    wb.save(response)
    return response

def download_excel_dailysheet(request):
    # Get the date filter from the query parameters
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter stocks based on the date_filter if it is provided
    if date_filter:
        stocks = DailySheet.objects.filter(date=date_filter).order_by('bird_type')
    elif start_date and end_date:
        stocks = DailySheet.objects.filter(date__range=[start_date, end_date]).order_by('date', 'bird_type')
    else:
        stocks = DailySheet.objects.all().order_by('date', 'bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given filter
        return HttpResponse("No records found for the given date range.", status=404)

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Stock"

    # Add the main heading
    ws.merge_cells('A1:H1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')  # Merge cells to center the heading
    ws['A2'] = "Opening Stock"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add a blank row and set the headers for date and day columns
    ws.append([""])
    ws.append(["Date:", "", "", "Day:", ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers and make the headers bold
    headers = ["DATE", "DAY", "BIRDS TYPE","NO.OF.BIRDS  C/O STOCK", "NO.OF.BIRS PURCHASE", "TOTAL BIRDS", "TOTAL C/O STOCK WEIGHT", "TOTAL PURCHASE WEIGHT", "TOTAL WEIGHT"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables for tracking merging and row positions
    previous_date = None
    row_start = 7  # Row number where the data starts

    # Add stock data to the Excel sheet with merged date and day cells
    for stock in stocks:
        current_row = ws.max_row + 1

        # Insert a gap row if the current stock's date is different from the previous date
        if stock.date != previous_date and previous_date is not None:
            ws.append([""] * len(headers))  # Insert an empty row
            current_row += 1

        row = [
            stock.date if stock.date != previous_date else "",  # Only show date if it's different from the previous row
            stock.day if stock.date != previous_date else "",  # Only show day if it's different from the previous row
            stock.bird_type,
            stock.number_of_birds_stock,
            stock.number_of_birds_purchase,
            stock.total_birds,
            stock.total_stock_weight,
            stock.total_purchase_weight,
            stock.total_weight
        ]
        ws.append(row)

        # Center align the date and day columns
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Merge date and day cells if necessary
        if stock.date != previous_date:
            if current_row > row_start:
                ws.merge_cells(start_row=row_start, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=row_start, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')
            row_start = current_row
            previous_date = stock.date

        # Apply styling to the rest of the row cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 3:  # Make bird_type (column 3) bold
                cell.font = Font(bold=True)

    # Merge the last set of date and day cells
    if row_start < ws.max_row:
        ws.merge_cells(start_row=row_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=row_start, start_column=2, end_row=ws.max_row, end_column=2)
        ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set up the filename
    if date_filter:
        filename = f"Opening_Stock_{date_filter}.xlsx"
    elif start_date and end_date:
        filename = f"Opening_Stock_{start_date}_to_{end_date}.xlsx"
    else:
        filename = "Opening_Stock.xlsx"
    
    # Set up the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def daily_sales(request):
    if request.method == 'POST':
        if 'delete' in request.POST:
            report_date = request.POST.get('report_date')
            print(f"Report date received: {report_date}")  # Debugging
            try:
                # Convert to date object
                report_date = datetime.strptime(report_date.strip(), "%Y-%m-%d").date()
                # Delete matching records
                deleted_count, _ = DailySales.objects.filter(date=report_date).delete()
                if deleted_count == 0:
                    print(f"No records found for date: {report_date}")  # Debugging
            except ValueError as e:
                return redirect('daily_sales')

            return redirect('daily_sales')
        
        date = request.POST.get('date')
        day = request.POST.get('day')

        bird_types = ['broiler', 'cc', 'original', 'quail']

        for bird_type in bird_types:
            # Convert inputs to Decimal, defaulting to 0 if empty
            live_weight = Decimal(request.POST.get(f'{bird_type}_live_weight', '0') or '0')
            curry_weight = Decimal(request.POST.get(f'{bird_type}_curry_weight', '0') or '0')
            day_rate = int(request.POST.get(f'{bird_type}_day_rate', '0') or '0')
            total_sales_amount = Decimal(request.POST.get(f'{bird_type}_total_sales_amount', '0') or '0')
            expense = Decimal(request.POST.get(f'{bird_type}_expense', '0') or '0')
            balance_cash = Decimal(request.POST.get(f'{bird_type}_balance_cash', '0') or '0')
            gpay = Decimal(request.POST.get(f'{bird_type}_gpay', '0') or '0')

            # Create a record only if any relevant field is provided
            if any([live_weight, curry_weight, day_rate, total_sales_amount, expense, balance_cash, gpay]):
                DailySales.objects.create(
                    date=date,
                    day=day,
                    bird_type=bird_type.capitalize(),
                    live_weight=live_weight,
                    curry_weight=curry_weight,
                    day_rate=day_rate,
                    total_sales_amount=total_sales_amount,
                    expense=expense,
                    balance_cash=balance_cash,
                    gpay=gpay
                )

        return redirect('daily_sales')  # Ensure this matches the URL name

    # Retrieve all existing stock entries
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Retrieve all distinct date and day combinations
    stocks = DailySales.objects.values('date', 'day').distinct().order_by('-date')

    if date_filter:
        stocks = stocks.filter(date=date_filter)
    elif start_date and end_date:
        stocks = stocks.filter(date__range=[start_date, end_date])

    return render(request, 'dailysales.html', {
        'stocks': stocks,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date
    })

def view_stock_dailysales(request, stock_date):
    # Query all stock entries for the given date and order by bird_type
    stocks = DailySales.objects.filter(date=stock_date).order_by('bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given date
        return HttpResponse("No records found for the given date.", status=404)

    # Get the day from the first stock entry (since all entries will have the same date and day)
    stock_day = stocks.first().day

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Sales"

    ws.merge_cells('A1:E1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add the "Closing Stock" heading, center it, and apply bold styling
    ws.merge_cells('A2:E2')  # Merge cells to center the heading
    ws['A2'] = "Daily Sales"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add the date and day, and make the heading bold
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stock_day])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers, add a "S.NO" column, and make the headers bold
    headers = ["S.NO", "BIRDS TYPE", "LIVE WEIGHT", "CURRY WEIGHT", "DAY RATE", "TOTAL SALES AMOUNT", "EXPENSE", "BALANCE CASH", "GPAY"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables to track the current group and serial number
    current_bird_type = None
    serial_number = 1

    for stock in stocks:
        if stock.bird_type != current_bird_type:
            # Start a new group if the bird_type changes
            current_bird_type = stock.bird_type

        # Append stock data for the current group, including the serial number
        row = [
            serial_number,
            stock.bird_type,
            stock.live_weight,
            stock.curry_weight,
            stock.day_rate,
            stock.total_sales_amount,
            stock.expense,
            stock.balance_cash,
            stock.gpay
        ]
        ws.append(row)

        # Apply bold styling to bird_type and center alignment to all cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:  # Make bird_type (column 2) bold
                cell.font = Font(bold=True)

        serial_number += 1

    # Set column widths for better visibility
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=DailySales_Report_{stock_date}.xlsx'

    wb.save(response)
    return response

def download_excel_dailysales(request):
    # Get the date filter from the query parameters
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter stocks based on the date_filter if it is provided
    if date_filter:
        stocks = DailySales.objects.filter(date=date_filter).order_by('bird_type')
    elif start_date and end_date:
        stocks = DailySales.objects.filter(date__range=[start_date, end_date]).order_by('date', 'bird_type')
    else:
        stocks = DailySales.objects.all().order_by('date', 'bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given filter
        return HttpResponse("No records found for the given date range.", status=404)

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Sales"

    # Add the main heading
    ws.merge_cells('A1:H1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')  # Merge cells to center the heading
    ws['A2'] = "Weekly Report"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add a blank row and set the headers for date and day columns
    ws.append([""])
    ws.append(["Date:", "", "", "Day:", ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers and make the headers bold
    headers = ["DATE", "DAY", "BIRDS TYPE", "LIVE WEIGHT", "CURRY WEIGHT", "DAY RATE", "TOTAL SALES AMOUNT", "EXPENSE", "BALANCE CASH", "GPAY"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables for tracking merging and row positions
    previous_date = None
    row_start = 7  # Row number where the data starts

    # Add stock data to the Excel sheet with merged date and day cells
    for stock in stocks:
        current_row = ws.max_row + 1

        # Insert a gap row if the current stock's date is different from the previous date
        if stock.date != previous_date and previous_date is not None:
            ws.append([""] * len(headers))  # Insert an empty row
            current_row += 1

        row = [
            stock.date if stock.date != previous_date else "",  # Only show date if it's different from the previous row
            stock.day if stock.date != previous_date else "",  # Only show day if it's different from the previous row
            stock.bird_type,
            stock.live_weight,
            stock.curry_weight,
            stock.day_rate,
            stock.total_sales_amount,
            stock.expense,
            stock.balance_cash,
            stock.gpay
        ]
        ws.append(row)

        # Center align the date and day columns
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Merge date and day cells if necessary
        if stock.date != previous_date:
            if current_row > row_start:
                ws.merge_cells(start_row=row_start, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=row_start, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')
            row_start = current_row
            previous_date = stock.date

        # Apply styling to the rest of the row cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 3:  # Make bird_type (column 3) bold
                cell.font = Font(bold=True)

    # Merge the last set of date and day cells
    if row_start < ws.max_row:
        ws.merge_cells(start_row=row_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=row_start, start_column=2, end_row=ws.max_row, end_column=2)
        ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set up the filename
    if date_filter:
        filename = f"Daily_Sales_{date_filter}.xlsx"
    elif start_date and end_date:
        filename = f"Daily_Sales_{start_date}_to_{end_date}.xlsx"
    else:
        filename = "Daily_Sales.xlsx"
    
    # Set up the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def weekly_report(request):
    if request.method == 'POST':
        if 'delete' in request.POST:
            report_date = request.POST.get('report_date')
            print(f"Report date received: {report_date}")  # Debugging
            try:
                # Convert to date object
                report_date = datetime.strptime(report_date.strip(), "%Y-%m-%d").date()
                # Delete matching records
                deleted_count, _ = WeeklyReport.objects.filter(date=report_date).delete()
                if deleted_count == 0:
                    print(f"No records found for date: {report_date}")  # Debugging
            except ValueError as e:
                return redirect('weekly_report')

            return redirect('weekly_report')
        
        date = request.POST.get('date')
        day = request.POST.get('day')
        
        bird_types = ['broiler', 'cc', 'original', 'quail']

        for bird_type in bird_types:
            number_of_birds = request.POST.get(f'number_of_birds_{bird_type}', 0)
            total_kilograms = request.POST.get(f'total_kilograms_{bird_type}', 0.0)
            average_weight = request.POST.get(f'average_weight_{bird_type}', 0.0)
            rate = request.POST.get(f'rate_{bird_type}', 0.0)
            total_amount = request.POST.get(f'total_amount_{bird_type}', 0.0)
            remarks = request.POST.get(f'remarks_{bird_type}', '')

            # Create a record only if any relevant field is provided
            if any([number_of_birds, total_kilograms, average_weight, rate, total_amount, remarks]):
                WeeklyReport.objects.create(
                    date=date,
                    day=day,
                    bird_type=bird_type.capitalize(),
                    number_of_birds=int(number_of_birds),
                    total_kilograms=float(total_kilograms),
                    average_weight=float(average_weight),
                    rate=float(rate),
                    total_amount=float(total_amount),
                    remarks=remarks
                )

        return redirect('weekly_report')
    
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Retrieve all distinct date and day combinations
    stocks = WeeklyReport.objects.values('date', 'day').distinct().order_by('-date')

    if date_filter:
        stocks = stocks.filter(date=date_filter)
    elif start_date and end_date:
        stocks = stocks.filter(date__range=[start_date, end_date])


    # Retrieve all distinct date and day combinations
    
    return render(request, 'weeklyreport.html', {'stocks': stocks, 'date_filter': date_filter, 'start_date': start_date,
        'end_date': end_date})

def view_stock_weeklyreport(request, stock_date):
    # Query all stock entries for the given date and order by bird_type
    stocks = WeeklyReport.objects.filter(date=stock_date).order_by('bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given date
        return HttpResponse("No records found for the given date.", status=404)

    # Get the day from the first stock entry (since all entries will have the same date and day)
    stock_day = stocks.first().day

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Purchase"

    ws.merge_cells('A1:E1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add the "Closing Stock" heading, center it, and apply bold styling
    ws.merge_cells('A2:E2')  # Merge cells to center the heading
    ws['A2'] = "Weekly Purchase"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add the date and day, and make the heading bold
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stock_day])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers, add a "S.NO" column, and make the headers bold
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "TOTAL KGS", "AVERAGE WEIGHT", "RATE", "TOTAL AMOUNT", "REMARKS"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables to track the current group and serial number
    current_bird_type = None
    serial_number = 1

    for stock in stocks:
        if stock.bird_type != current_bird_type:
            # Start a new group if the bird_type changes
            current_bird_type = stock.bird_type

        # Append stock data for the current group, including the serial number
        row = [
            serial_number,
            stock.bird_type,
            stock.number_of_birds,
            stock.total_kilograms,
            stock.average_weight,
            stock.rate,
            stock.total_amount,
            stock.remarks
        ]
        ws.append(row)

        # Apply bold styling to bird_type and center alignment to all cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:  # Make bird_type (column 2) bold
                cell.font = Font(bold=True)

        serial_number += 1

    # Set column widths for better visibility
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Weekly_Purchase_{stock_date}.xlsx'

    wb.save(response)
    return response

def download_excel(request):
    # Get the date filter from the query parameters
    date_filter = request.GET.get('date_filter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter stocks based on the date_filter if it is provided
    if date_filter:
        stocks = WeeklyReport.objects.filter(date=date_filter).order_by('bird_type')
    elif start_date and end_date:
        stocks = WeeklyReport.objects.filter(date__range=[start_date, end_date]).order_by('date', 'bird_type')
    else:
        stocks = WeeklyReport.objects.all().order_by('date', 'bird_type')
    
    if not stocks.exists():
        # Handle the case where no stock records exist for the given filter
        return HttpResponse("No records found for the given date range.", status=404)

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    # Add the main heading
    ws.merge_cells('A1:H1')  # Merge cells to center the heading
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')  # Merge cells to center the heading
    ws['A2'] = "Weekly Purchase"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add a blank row and set the headers for date and day columns
    ws.append([""])
    ws.append(["Date:", "", "", "Day:", ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])  # Empty row for separation

    # Define headers and make the headers bold
    headers = ["DATE", "DAY", "BIRDS TYPE", "NO.OF.BIRDS", "TOTAL KGS", "AVERAGE WEIGHT", "RATE", "TOTAL AMOUNT", "REMARKS"]
    ws.append(headers)
    for cell in ws[6]:  # Row 6 contains the headers
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Initialize variables for tracking merging and row positions
    previous_date = None
    row_start = 7  # Row number where the data starts

    # Add stock data to the Excel sheet with merged date and day cells
    for stock in stocks:
        current_row = ws.max_row + 1

        # Insert a gap row if the current stock's date is different from the previous date
        if stock.date != previous_date and previous_date is not None:
            ws.append([""] * len(headers))  # Insert an empty row
            current_row += 1

        row = [
            stock.date if stock.date != previous_date else "",  # Only show date if it's different from the previous row
            stock.day if stock.date != previous_date else "",  # Only show day if it's different from the previous row
            stock.bird_type,
            stock.number_of_birds,
            stock.total_kilograms,
            stock.average_weight,
            stock.rate,
            stock.total_amount,
            stock.remarks
        ]
        ws.append(row)

        # Center align the date and day columns
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Merge date and day cells if necessary
        if stock.date != previous_date:
            if current_row > row_start:
                ws.merge_cells(start_row=row_start, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=row_start, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')
            row_start = current_row
            previous_date = stock.date

        # Apply styling to the rest of the row cells
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 3:  # Make bird_type (column 3) bold
                cell.font = Font(bold=True)

    # Merge the last set of date and day cells
    if row_start < ws.max_row:
        ws.merge_cells(start_row=row_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=row_start, start_column=2, end_row=ws.max_row, end_column=2)
        ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_start, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)  # Convert column index to letter
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Avoid merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set up the filename
    if date_filter:
        filename = f"weekly_purchase_{date_filter}.xlsx"
    elif start_date and end_date:
        filename = f"weekly_purchase_{start_date}_to_{end_date}.xlsx"
    else:
        filename = "weekly_purchase.xlsx"
    
    # Set up the HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save the workbook to the response
    wb.save(response)
    return response

# def create_combined_excel(stock_date):
#     wb = openpyxl.Workbook()
#     del wb['Sheet']  # Remove the default sheet

#     # Create Closing Stock Sheet
#     stocks = ClosingStock.objects.filter(date=stock_date).order_by('bird_type')
#     ws = wb.create_sheet(title="Closing Report")
#     ws.merge_cells('A1:E1')
#     ws['A1'] = "Farm Nutri Chicken"
#     ws['A1'].font = Font(size=16, bold=True)
#     ws['A1'].alignment = Alignment(horizontal='center')
#     ws.merge_cells('A2:E2')
#     ws['A2'] = "Closing Stock"
#     ws['A2'].font = Font(size=14, bold=True)
#     ws['A2'].alignment = Alignment(horizontal='center')
#     ws.append([""])
#     ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
#     ws['A4'].font = Font(bold=True)
#     ws['D4'].font = Font(bold=True)
#     ws.append([""])
#     headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "NO.OF.KGS", "MORTALITY"]
#     ws.append(headers)
#     for cell in ws[6]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')
#     serial_number = 1
#     for stock in stocks:
#         row = [
#             serial_number,
#             stock.bird_type,
#             stock.no_of_birds,
#             stock.no_of_kgs,
#             stock.mortality,
#         ]
#         ws.append(row)
#         for i, value in enumerate(row, 1):
#             cell = ws.cell(row=ws.max_row, column=i)
#             cell.alignment = Alignment(horizontal='center')
#             if i == 2:
#                 cell.font = Font(bold=True)
#         serial_number += 1
#     for col_num, col_cells in enumerate(ws.columns, 1):
#         max_length = 0
#         column_letter = get_column_letter(col_num)
#         for cell in col_cells:
#             if not isinstance(cell, openpyxl.cell.cell.MergedCell):
#                 try:
#                     if cell.value and len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width

#     # Create Daily Sheet
#     stocks = DailySheet.objects.filter(date=stock_date).order_by('bird_type')
#     ws = wb.create_sheet(title="Daily Sheet")
#     ws.merge_cells('A1:H1')
#     ws['A1'] = "Farm Nutri Chicken"
#     ws['A1'].font = Font(size=16, bold=True)
#     ws['A1'].alignment = Alignment(horizontal='center')
#     ws.merge_cells('A2:H2')
#     ws['A2'] = "Daily Sheet"
#     ws['A2'].font = Font(size=14, bold=True)
#     ws['A2'].alignment = Alignment(horizontal='center')
#     ws.append([""])
#     ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
#     ws['A4'].font = Font(bold=True)
#     ws['D4'].font = Font(bold=True)
#     ws.append([""])
#     headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS C/O STOCK", "NO.OF.BIRDS PURCHASE", "TOTAL BIRDS", "TOTAL C/O STOCK WEIGHT", "TOTAL PURCHASE WEIGHT", "TOTAL WEIGHT"]
#     ws.append(headers)
#     for cell in ws[6]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')
#     serial_number = 1
#     for stock in stocks:
#         row = [
#             serial_number,
#             stock.bird_type,
#             stock.number_of_birds_stock,
#             stock.number_of_birds_purchase,
#             stock.total_birds,
#             stock.total_stock_weight,
#             stock.total_purchase_weight,
#             stock.total_weight
#         ]
#         ws.append(row)
#         for i, value in enumerate(row, 1):
#             cell = ws.cell(row=ws.max_row, column=i)
#             cell.alignment = Alignment(horizontal='center')
#             if i == 2:
#                 cell.font = Font(bold=True)
#         serial_number += 1
#     for col_num, col_cells in enumerate(ws.columns, 1):
#         max_length = 0
#         column_letter = get_column_letter(col_num)
#         for cell in col_cells:
#             if not isinstance(cell, openpyxl.cell.cell.MergedCell):
#                 try:
#                     if cell.value and len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width

#     # Create Daily Sales Sheet
#     stocks = DailySales.objects.filter(date=stock_date).order_by('bird_type')
#     ws = wb.create_sheet(title="Daily Sales")
#     ws.merge_cells('A1:I1')
#     ws['A1'] = "Farm Nutri Chicken"
#     ws['A1'].font = Font(size=16, bold=True)
#     ws['A1'].alignment = Alignment(horizontal='center')
#     ws.merge_cells('A2:I2')
#     ws['A2'] = "Daily Sales"
#     ws['A2'].font = Font(size=14, bold=True)
#     ws['A2'].alignment = Alignment(horizontal='center')
#     ws.append([""])
#     ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
#     ws['A4'].font = Font(bold=True)
#     ws['D4'].font = Font(bold=True)
#     ws.append([""])
#     headers = ["S.NO", "BIRDS TYPE", "LIVE WEIGHT", "CURRY WEIGHT", "DAY RATE", "TOTAL SALES AMOUNT", "EXPENSE", "BALANCE CASH", "GPAY"]
#     ws.append(headers)
#     for cell in ws[6]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')
#     serial_number = 1
#     for stock in stocks:
#         row = [
#             serial_number,
#             stock.bird_type,
#             stock.live_weight,
#             stock.curry_weight,
#             stock.day_rate,
#             stock.total_sales_amount,
#             stock.expense,
#             stock.balance_cash,
#             stock.gpay
#         ]
#         ws.append(row)
#         for i, value in enumerate(row, 1):
#             cell = ws.cell(row=ws.max_row, column=i)
#             cell.alignment = Alignment(horizontal='center')
#             if i == 2:
#                 cell.font = Font(bold=True)
#         serial_number += 1
#     for col_num, col_cells in enumerate(ws.columns, 1):
#         max_length = 0
#         column_letter = get_column_letter(col_num)
#         for cell in col_cells:
#             if not isinstance(cell, openpyxl.cell.cell.MergedCell):
#                 try:
#                     if cell.value and len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width

#     # Create Weekly Report Sheet
#     stocks = WeeklyReport.objects.filter(date=stock_date).order_by('bird_type')
#     ws = wb.create_sheet(title="Weekly Report")
#     ws.merge_cells('A1:H1')
#     ws['A1'] = "Farm Nutri Chicken"
#     ws['A1'].font = Font(size=16, bold=True)
#     ws['A1'].alignment = Alignment(horizontal='center')
#     ws.merge_cells('A2:H2')
#     ws['A2'] = "Weekly Report"
#     ws['A2'].font = Font(size=14, bold=True)
#     ws['A2'].alignment = Alignment(horizontal='center')
#     ws.append([""])
#     ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
#     ws['A4'].font = Font(bold=True)
#     ws['D4'].font = Font(bold=True)
#     ws.append([""])
#     headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "TOTAL KGS", "AVERAGE WEIGHT", "RATE", "TOTAL AMOUNT", "REMARKS"]
#     ws.append(headers)
#     for cell in ws[6]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')
#     serial_number = 1
#     for stock in stocks:
#         row = [
#             serial_number,
#             stock.bird_type,
#             stock.number_of_birds,
#             stock.total_kilograms,
#             stock.average_weight,
#             stock.rate,
#             stock.total_amount,
#             stock.remarks
#         ]
#         ws.append(row)
#         for i, value in enumerate(row, 1):
#             cell = ws.cell(row=ws.max_row, column=i)
#             cell.alignment = Alignment(horizontal='center')
#             if i == 2:
#                 cell.font = Font(bold=True)
#         serial_number += 1
#     for col_num, col_cells in enumerate(ws.columns, 1):
#         max_length = 0
#         column_letter = get_column_letter(col_num)
#         for cell in col_cells:
#             if not isinstance(cell, openpyxl.cell.cell.MergedCell):
#                 try:
#                     if cell.value and len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width

#     # Save the workbook to a temporary file
#     file_name = f"Combined_Report_{stock_date}.xlsx"
#     wb.save(file_name)
#     return file_name
