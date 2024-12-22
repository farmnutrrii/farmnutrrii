import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.core.mail import EmailMessage
from .models import ClosingStock, DailySheet, DailySales, WeeklyReport

def create_combined_excel(stock_date):
    wb = openpyxl.Workbook()
    del wb['Sheet']  # Remove the default sheet

    # Create Closing Stock Sheet
    stocks = ClosingStock.objects.filter(date=stock_date).order_by('bird_type')
    ws = wb.create_sheet(title="Closing Report")
    ws.merge_cells('A1:E1')
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:E2')
    ws['A2'] = "Closing Stock"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "NO.OF.KGS", "MORTALITY"]
    ws.append(headers)
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    serial_number = 1
    for stock in stocks:
        row = [
            serial_number,
            stock.bird_type,
            stock.no_of_birds,
            stock.no_of_kgs,
            stock.mortality,
        ]
        ws.append(row)
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:
                cell.font = Font(bold=True)
        serial_number += 1
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create Daily Sheet
    stocks = DailySheet.objects.filter(date=stock_date).order_by('bird_type')
    ws = wb.create_sheet(title="Daily Sheet")
    ws.merge_cells('A1:H1')
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:H2')
    ws['A2'] = "Opening Stock"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS C/O STOCK", "NO.OF.BIRDS PURCHASE", "TOTAL BIRDS", "TOTAL C/O STOCK WEIGHT", "TOTAL PURCHASE WEIGHT", "TOTAL WEIGHT"]
    ws.append(headers)
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    serial_number = 1
    for stock in stocks:
        row = [
            serial_number,
            stock.bird_type,
            stock.number_of_birds_stock,
            stock.number_of_birds_purchase,
            stock.total_birds,
            stock.total_stock_weight,
            stock.total_purchase_weight,
            stock.total_weight
        ]
        ws.append(row)
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:
                cell.font = Font(bold=True)
        serial_number += 1
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create Daily Sales Sheet
    stocks = DailySales.objects.filter(date=stock_date).order_by('bird_type')
    ws = wb.create_sheet(title="Daily Sales")
    ws.merge_cells('A1:I1')
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')
    ws['A2'] = "Daily Sales"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])
    headers = ["S.NO", "BIRDS TYPE", "LIVE WEIGHT", "CURRY WEIGHT", "DAY RATE", "TOTAL SALES AMOUNT", "EXPENSE", "BALANCE CASH", "GPAY"]
    ws.append(headers)
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    serial_number = 1
    for stock in stocks:
        row = [
            serial_number,
            stock.bird_type,
            stock.live_weight,
            stock.curry_weight,
            stock.day_rate,
            stock.total_sales_amount,
            stock.expense,
            stock.balance_cash,
            stock.gpay
        ]
        ws.append(row)
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:
                cell.font = Font(bold=True)
        serial_number += 1
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create Weekly Report Sheet
    stocks = WeeklyReport.objects.filter(date=stock_date).order_by('bird_type')
    ws = wb.create_sheet(title="Weekly Purchase")
    ws.merge_cells('A1:H1')
    ws['A1'] = "Farm Nutri Chicken"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:H2')
    ws['A2'] = "Weekly Report"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([""])
    ws.append(["Date:", stock_date, "", "Day:", stocks.first().day if stocks.exists() else ""])
    ws['A4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws.append([""])
    headers = ["S.NO", "BIRDS TYPE", "NO.OF.BIRDS", "TOTAL KGS", "AVERAGE WEIGHT", "RATE", "TOTAL AMOUNT", "REMARKS"]
    ws.append(headers)
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    serial_number = 1
    for stock in stocks:
        row = [
            serial_number,
            stock.bird_type,
            stock.number_of_birds,
            stock.total_kilograms,
            stock.average_weight,
            stock.rate,
            stock.total_amount,
            stock.remarks
        ]
        ws.append(row)
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = Alignment(horizontal='center')
            if i == 2:
                cell.font = Font(bold=True)
        serial_number += 1
    for col_num, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to a temporary file
    file_name = f"Combined_Report_{stock_date}.xlsx"
    wb.save(file_name)
    return file_name

def send_email_with_report(stock_date):
    file_name = create_combined_excel(stock_date)

    email = EmailMessage(
        subject=f"Daily Stock Report for {stock_date}",
        body="Please find the attached stock report.",
        from_email='your-email@example.com',
        to=['recipient@example.com'],  # Add more recipients as needed
    )
    email.attach_file(file_name)
    email.send()

def view_stock_combined_report(request, stock_date):
    send_email_with_report(stock_date)
    return HttpResponse("Combined report sent successfully.")
